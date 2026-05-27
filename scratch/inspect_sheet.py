import openpyxl

excel_path = "c:\\AI\\TimeTracker\\DAILY Tasking Status Report ( 2026_05_25 - 2026_05_29 ).xlsx"
wb = openpyxl.load_workbook(excel_path)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: {name}, Dimensions: {ws.dimensions}")
    # Print headers (row 1)
    headers = [cell.value for cell in ws[1]]
    print("Headers:", headers)
    # Print first few rows
    for r in range(2, min(15, ws.max_row + 1)):
        row_vals = [cell.value for cell in ws[r]]
        if any(row_vals):
            print(f"Row {r}:", row_vals[:18])
